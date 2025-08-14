import os, re, time, subprocess, sys
from datetime import datetime

import pandas as pd
import yfinance as yf

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils.cell import get_column_letter

# ========= SETTINGS =========
MASTER_XLSX = "SKV Sheet_Updated PM.xlsx"    # You edit this file
OUTPUT_CSV  = "SKV Sheet-1-Updated.csv"      # Power Query / GitHub raw
THRESHOLD   = 0.025                           # 2.5% for green/red highlights
SLEEP_SEC   = 0.3                             # polite delay for Yahoo
AUTO_PUSH   = True                            # set False to skip git push

# ========= HELPERS =========
def log(msg): print(msg, flush=True)

def run(cmd, cwd=None):
    """Run a shell command; return (rc, stdout, stderr)."""
    try:
        p = subprocess.Popen(
            cmd, cwd=cwd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=True
        )
        out, err = p.communicate()
        return p.returncode, out.decode(errors="ignore"), err.decode(errors="ignore")
    except Exception as e:
        return 1, "", str(e)

def git_available():
    rc, out, _ = run("git --version")
    return rc == 0

def git_commit_push(files, message):
    if not AUTO_PUSH:
        log("ℹ️ AUTO_PUSH=False — skipping git add/commit/push.")
        return
    if not git_available():
        log("⚠️ Git not found on PATH; skipping auto-push. Install Git to enable.")
        return
    # Ensure we're inside a git repo
    rc, out, err = run("git rev-parse --is-inside-work-tree")
    if rc != 0:
        log("⚠️ Not in a Git repository; skipping auto-push.")
        return

    # Add files
    for f in files:
        run(f'git add "{f}"')

    # Commit (allow empty = no changes)
    rc, out, err = run(f'git commit -m "{message}"')
    if rc != 0 and "nothing to commit" in (out + err).lower():
        log("ℹ️ No changes to commit.")
    else:
        log(out or err)

    # Push
    rc, out, err = run("git push")
    if rc != 0:
        log(f"⚠️ git push failed:\n{err or out}")
    else:
        log("🚀 Pushed to remote.")

def ensure_columns(df, cols):
    for c in cols:
        if c not in df.columns:
            df[c] = None
    return df

def clean_symbol(sym):
    if not isinstance(sym, str) or sym.strip() == "":
        return None
    s = sym.strip().upper()
    s = re.sub(r"^\$+", "", s)
    s = s.replace("_", "-")
    s = re.sub(r"[^A-Z0-9\-\.]", "", s)
    if not s.endswith(".NS"):
        s += ".NS"
    return s

# ========= VALIDATE FILE =========
if not os.path.exists(MASTER_XLSX):
    raise FileNotFoundError(f"{MASTER_XLSX} not found in repo!")

# ========= LOAD =========
df = pd.read_excel(MASTER_XLSX, engine="openpyxl")

# Keep schema stable (add missing columns if user added only a few)
required_cols = [
    "Stock Name","Time Frame","Zone","Entry Price","Stop Loss","Legout Date",
    "Validation Issue","Zone Perfection","Entry Date","Status","Diff","Qty",
    "Tgt","Yahoo Symbol","Last Close Price"
]
df = ensure_columns(df, required_cols)

# Fill Yahoo Symbol for new/blank rows based on Stock Name
df["Yahoo Symbol"] = df["Yahoo Symbol"].where(df["Yahoo Symbol"].notna(), df["Stock Name"])
df["Yahoo Symbol"] = df["Yahoo Symbol"].apply(clean_symbol)

# ========= FETCH LATEST PRICES =========
new_prices = {}
failed = []

symbols = list(df["Yahoo Symbol"]) if "Yahoo Symbol" in df.columns else []
for symbol in symbols:
    if pd.isna(symbol) or not isinstance(symbol, str) or symbol.strip() == "":
        failed.append(symbol)
        continue
    try:
        t = yf.Ticker(symbol)
        hist = t.history(period="1d")
        if hist.empty:
            hist = t.history(period="5d")
        if not hist.empty:
            price = float(hist["Close"].dropna().iloc[-1])
            new_prices[symbol] = round(price, 2)
        else:
            failed.append(symbol)
    except Exception:
        failed.append(symbol)
    time.sleep(SLEEP_SEC)

def set_price(row):
    sym = row.get("Yahoo Symbol")
    if isinstance(sym, str) and sym in new_prices:
        return new_prices[sym]
    if isinstance(sym, str) and sym in failed:
        return "Not available"
    # keep existing if neither fetched nor failed
    return row.get("Last Close Price")

df["Last Close Price"] = df.apply(set_price, axis=1)

# ========= DIFF % helper =========
def diff_pct(row):
    entry = row.get("Entry Price")
    last  = row.get("Last Close Price")
    try:
        entry = float(entry)
        last  = float(last)
        if entry and entry != 0:
            return round((last - entry) / entry * 100, 2)
    except Exception:
        return None
    return None

df["Diff %"] = df.apply(diff_pct, axis=1)

# ========= SAVE (Excel + CSV) =========
df.to_excel(MASTER_XLSX, index=False)
df.to_csv(OUTPUT_CSV, index=False)

# ========= CONDITIONAL FORMATTING (±2.5%) =========
wb = load_workbook(MASTER_XLSX)
ws = wb.active

# Map headers -> column letters
headers = {cell.value: cell.column for cell in ws[1] if cell.value}
def col_letter(col_name):
    idx = headers.get(col_name)
    return get_column_letter(idx) if idx else None

col_entry = col_letter("Entry Price")
col_last  = col_letter("Last Close Price")

if col_entry and col_last and ws.max_row >= 2:
    start_row = 2
    end_row = ws.max_row
    end_col_letter = get_column_letter(ws.max_column)
    full_range = f"A{start_row}:{end_col_letter}{end_row}"

    # Use a relative formula anchored on the first data row; PQ copies it per-row
    green_formula = f"=AND(ISNUMBER(${col_last}{start_row}),ISNUMBER(${col_entry}{start_row}),${col_last}{start_row}>={col_entry}{start_row}*(1+{THRESHOLD}))"
    red_formula   = f"=AND(ISNUMBER(${col_last}{start_row}),ISNUMBER(${col_entry}{start_row}),${col_last}{start_row}<={col_entry}{start_row}*(1-{THRESHOLD}))"

    green_rule = FormulaRule(formula=[green_formula],
                             fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                             stopIfTrue=True)
    red_rule   = FormulaRule(formula=[red_formula],
                             fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                             stopIfTrue=True)

    # Clear old rules (optional but avoids stacking duplicates)
    ws.conditional_formatting._cf_rules = {}

    ws.conditional_formatting.add(full_range, green_rule)
    ws.conditional_formatting.add(full_range, red_rule)

wb.save(MASTER_XLSX)

# ========= AUTO COMMIT & PUSH =========
commit_msg = f'Auto-update stock prices @ {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
git_commit_push([MASTER_XLSX, OUTPUT_CSV], commit_msg)

# ========= LOG =========
log(f"✅ Prices updated at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
log(f"📂 Excel updated: {MASTER_XLSX}")
log(f"📂 CSV updated:   {OUTPUT_CSV}")

if failed:
    log("\n⚠️ Not available:")
    for s in sorted(set([x for x in failed if x])):
        log(f" - {s}")
